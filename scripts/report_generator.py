from openpyxl import Workbook
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def write_dataframe_to_sheet(sheet, dataframe):
    # Write column headers
    for col_num, column_name in enumerate(dataframe.columns, 1):
        cell = sheet.cell(row=1, column=col_num, value=column_name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
        cell.alignment = Alignment(horizontal="center")

    # Write dataframe rows
    for row_num, row_data in enumerate(dataframe.values, 2):
        for col_num, value in enumerate(row_data, 1):
            sheet.cell(row=row_num, column=col_num, value=value)

    # Auto-adjust column width
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        sheet.column_dimensions[column_letter].width = max_length + 2

    # Freeze header row
    sheet.freeze_panes = "A2"


def generate_excel_report(
    kpis,
    product_summary,
    top10,
    bottom10,
    category_summary,
    region_summary,
    state_summary,
    city_summary,
    channel_summary,
    monthly_summary
):
    """Generate Excel Business Report"""

    wb = Workbook()

    # =========================
    # Summary Sheet
    # =========================
    ws = wb.active
    ws.title = "Summary"

    ws["A1"] = "Sales Automation Report"
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A3"] = "KPI"
    ws["B3"] = "Value"

    for cell in ["A3", "B3"]:
        ws[cell].font = Font(bold=True, color="FFFFFF")
        ws[cell].fill = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
        ws[cell].alignment = Alignment(horizontal="center")

    row = 4
    for key, value in kpis.items():
        ws[f"A{row}"] = key
        ws[f"B{row}"] = value
        row += 1

    # Format summary sheet width
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18

    # Freeze summary header
    ws.freeze_panes = "A4"

    # =========================
    # Create analysis sheets
    # =========================
    product_ws = wb.create_sheet("Product Analysis")
    top10_ws = wb.create_sheet("Top 10 Products")
    bottom10_ws = wb.create_sheet("Bottom 10 Products")
    category_ws = wb.create_sheet("Category Analysis")
    region_ws = wb.create_sheet("Region Analysis")
    state_ws = wb.create_sheet("State Analysis")
    city_ws = wb.create_sheet("City Analysis")
    channel_ws = wb.create_sheet("Channel Analysis")
    monthly_ws = wb.create_sheet("Monthly Trend")

    # =========================
    # Write dataframes
    # =========================
    write_dataframe_to_sheet(product_ws, product_summary)
    write_dataframe_to_sheet(top10_ws, top10)
    write_dataframe_to_sheet(bottom10_ws, bottom10)
    write_dataframe_to_sheet(category_ws, category_summary)
    write_dataframe_to_sheet(region_ws, region_summary)
    write_dataframe_to_sheet(state_ws, state_summary)
    write_dataframe_to_sheet(city_ws, city_summary)
    write_dataframe_to_sheet(channel_ws, channel_summary)
    write_dataframe_to_sheet(monthly_ws, monthly_summary)
    # =========================
    # Apply number formatting
    # =========================
    revenue_sheets = [
    product_ws, top10_ws, bottom10_ws,
    category_ws, region_ws, state_ws,
    city_ws, channel_ws, monthly_ws
]

    for sheet in revenue_sheets:
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'

    for row_cells in ws.iter_rows(min_row=4, max_col=2):
        row_cells[1].number_format = '#,##0.00'

    # Save workbook
    wb.save("reports/Sales_Report.xlsx")